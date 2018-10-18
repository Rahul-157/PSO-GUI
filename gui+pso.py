from tkinter import *
from tkinter import filedialog
import numpy as np
from openpyxl import *


ent=[0,0,0,0,0,0,0,0,0,0]
#################MENU##############################
def new(event="<'Button-1'>"):
    temp_win = Toplevel(window)
    temp_win.title("New")
    temp_win.transient(window)
    temp_win.focus_set()
    temp_win.minsize(250, 100)
    temp_win.maxsize(250, 100)
    temp_win.geometry("250x100+450+250")
    temp_win.wm_iconbitmap('excel_icon.ico')
    name_lbl=Label(temp_win,text="Enter File name where data to be saved ")
    name_lbl.focus_set()
    name_lbl.pack()
    name_entry=Entry(temp_win)
    name_entry.pack()
    
    def create(event='<Button-1>'):
         Excel_file.set(name_entry.get()+".xlsx")
         temp_win.destroy()
         
    name_entry.bind('<Return>',create)     
    create_btn=Button(temp_win,command=create,text="Create")
    create_btn.pack()
    create_btn.place(height=25,width=100,x=75,y=50)


def exit_win(event="<'Button-1'>"):
    window.destroy()
########################MENU ENDS################

##############INITIALIZING GUI###################
def create_gui():
    lbl1=Label(window,text="Select Variables",relief="sunken",font='Comic 10 bold',bg="white")
    lbl1.pack()
    lbl1.place(x=150,y=50,height=40,width=180)
    
    option=OptionMenu(window,var,1,2,3,4,5,command=start)
    option.pack()
    option.place(x=350,y=50,height=40)
    var.set(2); #default selection
    start('<Button-1>');
    
    lbl4=Label(window,relief="groove")
    lbl4.pack()
    lbl4.place(x=150,y=355,height=100,width=755)
    
    lbl_c1=Label(window,text="C1",relief="raised",font='Comic 10 bold')
    lbl_c1.pack()
    lbl_c1.place(x=205,y=360)
    
    lbl_c2=Label(window,text="C2",relief="raised",font='Comic 10 bold')
    lbl_c2.pack()
    lbl_c2.place(x=360,y=360)
    
    lbl_wmax=Label(window,text="Wmax",relief="raised",font='Comic 10 bold')
    lbl_wmax.pack()
    lbl_wmax.place(x=505,y=360)
    
    lbl_wmin=Label(window,text="Wmin",relief="raised",font='Comic 10 bold')
    lbl_wmin.pack()
    lbl_wmin.place(x=662,y=360)
    
    lbl_np=Label(window,text="Num_Particles",relief="raised",font='Comic 10 bold')
    lbl_np.pack()
    lbl_np.place(x=788,y=360)
    
    lbl_fact=Label(window,text="Factors",relief="sunken",font='Comic 10 bold',bg="white")
    lbl_fact.pack()
    lbl_fact.place(x=10,y=400,height=20,width=100)
    
    lbl_eqn=Label(window,text="Equation",relief="sunken",font='Comic 10 bold',bg="white")
    lbl_eqn.pack()
    lbl_eqn.place(x=10,y=510,height=20,width=100)
    global ent_c1
    global ent_c2
    global ent_wmin
    global ent_wmax
    global ent_eqn
    global ent_np
    ent_c1=Entry(window,textvariable=c1)
    ent_c1.pack()
    ent_c1.place(x=155,y=400)
    
    ent_c2=Entry(window,textvariable=c2)
    ent_c2.pack()
    ent_c2.place(x=310,y=400)
    
    ent_wmax=Entry(window,textvariable=wmax)
    ent_wmax.pack()
    ent_wmax.place(x=465,y=400)
   
    ent_wmin=Entry(window,textvariable=wmin)
    ent_wmin.pack()
    ent_wmin.place(x=620,y=400)
    
    ent_np=Entry(window,textvariable=num_p)
    ent_np.pack()
    ent_np.place(x=775,y=400)

    ent_eqn=Entry(window,textvariable=eqn,font='fixedsys 16')
    ent_eqn.pack()
    ent_eqn.place(x=150,y=500,height=40,width=750)

    menubar=Menu(window)
    filemenu=Menu(menubar,tearoff=0)
    filemenu.add_command(label="New Ctrl + N",command=new)
    filemenu.add_separator()
    filemenu.add_command(label="Exit Ctrl + Q",command=exit_win)
    menubar.add_cascade(label="File",menu=filemenu)
    window.config(menu=menubar)
    window.bind('<Control-n>',new)
    window.bind('<Control-q>',exit_win)
    
    global t
    t=Text(window,height=41,width=55,yscrollcommand=scrollbar.set,xscrollcommand=scrollbarx.set)
    t.pack()
    t.place(x=905)
    t.insert(END,"\t     RESULT IN EACH ITERATION\n")
    t.config(state=DISABLED,wrap=NONE)
    scrollbar.config(command=t.yview)
    scrollbarx.config(command=t.xview)
    
def start(event):

    lbl2=Label(window,relief="groove")
    lbl2.pack()
    lbl2.place(x=150,y=150,height=200,width=755)
    lbl4=Label(window,text="Minimum",relief="sunken",font='Comic 10 bold',bg="white")
    lbl4.pack()
    lbl4.place(x=10,y=200,height=20,width=100)
    lbl5=Label(window,text="Maximum",relief="sunken",font='Comic 10 bold',bg="white")
    lbl5.pack()
    lbl5.place(x=10,y=300,height=20,width=100)
    for i in range(0,var.get()):
        lbl3=Label(window,text=str(Vars[i]),font='Comic 12 bold',relief='raised')
        lbl3.pack()
        lbl3.place(x=210+(i*155),y=155)
        for j in range(2):
            ent[i*2+j]=Entry(window)
            ent[i*2+j].pack()
            ent[i*2+j].place(x=(i+1)*155,y=(j+1)*100+100)

#####################GUI ENDS############################











####################FINISH MSG############################
            
def showmsg():
    temp_win = Toplevel(window)
    temp_win.title("SUCCESS!")
    temp_win.transient(window)
    temp_win.focus_set()
    temp_win.minsize(250, 50)
    temp_win.maxsize(250, 50)
    temp_win.geometry("250x100+450+250")
    temp_win.wm_iconbitmap('success.ico')
    name_lbl=Label(temp_win,text=str("DATA SAVED IN "+Excel_file.get()),font='fixedsys 16')
    name_lbl.focus_set()
    name_lbl.pack(fill=X)
    name_lbl.place(y=20)
    
###########################################################    









    
###########################################################





    
from numpy import *
from random import *
from py_expression_eval import Parser







########################PSO##############################
def pso():


    
    ##################Loading Excel File#####################
    global wb
    global ws
    wb=Workbook()
    wb.save(str(".\\"+Excel_file.get()));
    wb.close()
    wb=load_workbook(Excel_file.get())
    ws=wb.active
    ##################Loading Excel File Done################



    
    t.delete(1.0,END)
    t.insert(END,"\t   RESULT AT EACH ITERATION\n\n")
    window.update_idletasks()
    parser=Parser()
    n=var.get()
    itermax=500
    Pmin= 0
    Pmax= 1000
    c1=float(ent_c1.get())
    c2=float(ent_c2.get())
    pop=int(ent_np.get())
    error=zeros((itermax,1),dtype=float)
    global P
    global pbest
    global gbest
    global fit
    global v
    global w
    global Lp
    global Lg
    global fmin
    global fmax
    global Vmax_n
    global gbt
    global L
    global vmax
    P=zeros((pop,1),dtype=float)
    fmax=zeros((n,1),dtype=int)
    fmin=zeros((n,1),dtype=int)
    pbest=zeros((pop,1),dtype=float)
    gbest=0
    fit = zeros((pop,1),dtype=float)
    v = zeros((pop,n),dtype=float)
    vmax = zeros((n,1),dtype=float)
    wmax=float(ent_wmax.get())
    wmin=float(ent_wmin.get())
    Lp = zeros((pop,n),dtype=float)
    err = zeros((pop,1),dtype=float)
    Lg = zeros((n,1),dtype=float)
    for i in range (n):
        fmax[i]=ent[2*i+1].get()

        
    for i in range (n):
        fmin[i]=ent[2*i].get()
        
    Vmax_n = [100,100,100,100,100]
    gbt=0
    L = zeros((pop,n),dtype=float)
    parser.ops1['sin'] = sin
    parser.ops1['cos'] = cos
    parser.ops1['exp'] = exp
    parser.ops1['sqrt'] = sqrt
    parser.ops2['^'] = power
    pi=3.14159
    t.config(state=NORMAL)


    for j in range(pop):
        for k in range(n):
            L[j,k] = (randint(fmin[k],fmax[k]))+randint(1,100)/100;
    print(L)
    for i in range(n):
        vmax[i] = (fmax[i]-fmin[i]/Vmax_n[i]);
    if(n==1):
        print('x          fcn value ')
        t.insert(END,'x        fcn value\n')
        ws.cell(row=1,column=1).value="X"
        ws.cell(row=1,column=2).value="F(x)"
        ws.cell(row=1,column=3).value="ITERATION"
    elif(n==2):
        print('x      y    fcn value ')
        t.insert(END,'x     y     fcn value\n')
        ws.cell(row=1,column=1).value="X"
        ws.cell(row=1,column=2).value="Y"
        ws.cell(row=1,column=3).value="F(x,y)"
        ws.cell(row=1,column=4).value="ITERATION"
    elif(n==3):
        print('x      y      z      fcn value ')
        t.insert(END,'x     y      z      fcn value\n')
        ws.cell(row=1,column=1).value="X"
        ws.cell(row=1,column=2).value="Y"
        ws.cell(row=1,column=3).value="Z"
        ws.cell(row=1,column=4).value="F(x,y,z)"
        ws.cell(row=1,column=5).value="ITERATION"
    elif(n==4):
        print('x      y      z      r     fcn value \n')
        t.insert(END,'x     y      z      r      fcn value')
        ws.cell(row=1,column=1).value="X"
        ws.cell(row=1,column=2).value="Y"
        ws.cell(row=1,column=3).value="Z"
        ws.cell(row=1,column=4).value="R"
        ws.cell(row=1,column=5).value="F(x,y,z,r)"
        ws.cell(row=1,column=6).value="ITERATION"
    elif(n==5):
        print('x      y      z      r     s      fcn value \n')
        t.insert(END,'x     y      z      r      s      fcn value')
        ws.cell(row=1,column=1).value="X"
        ws.cell(row=1,column=2).value="Y"
        ws.cell(row=1,column=3).value="Z"
        ws.cell(row=1,column=4).value="R"
        ws.cell(row=1,column=5).value="S"
        ws.cell(row=1,column=6).value="F(x,y,z,r,s)"
        ws.cell(row=1,column=7).value="ITERATION"
    window.update_idletasks()

    for iteration in range(1,itermax+1):
        for i in range (pop):
            for j in range (n):
                if L[i][j]<=fmin[j]:
                    L[i][j]=float(fmin[j]);
                if L[i][j]>=fmax[j]:
                    L[i][j]=float(fmax[j]);
        if(n==1):
            X=array(L[:,0])
            P=parser.parse(ent_eqn.get()).evaluate({'x':X,'pi':pi})
        elif(n==2):
            X=array(L[:,0])
            Y=array(L[:,1])
            P=parser.parse(ent_eqn.get()).evaluate({'x':X,'y':Y,'pi':pi})
        elif(n==3):
            X=array(L[:,0])
            Y=array(L[:,1])
            Z=array(L[:,2])
            P=parser.parse(ent_eqn.get()).evaluate({'x':X,'y':Y,'z':Z,'pi':pi})
        elif(n==4):
            X=array(L[:,0])
            Y=array(L[:,1])
            Z=array(L[:,2])
            R=array(L[:,3])
            P=parser.parse(ent_eqn.get()).evaluate({'x':X,'y':Y,'z':Z,'r':R,'pi':pi})
        elif(n==5):
            X=array(L[:,0])
            Y=array(L[:,1])
            Z=array(L[:,2])
            R=array(L[:,3])
            S=array(L[:,4])
            P=parser.parse(ent_eqn.get()).evaluate({'x':X,'y':Y,'z':Z,'r':R,'s':S,'pi':pi})
        
        
        for i in range (pop):
            if P[i]<=Pmin:
                P[i]=Pmin
            elif P[i]>=Pmax:
                P[i]=Pmax

        for i in range(pop):
            fit[i]=1/(1+abs(P[i]))
            
            if fit[i]>=pbest[i]:
                pbest[i]=fit[i]
                Lp[i][0]=L[i][0]
                Lp[i][1]=L[i][1]

        if max(pbest)>=gbest:
            gbest=max(pbest)
        for i in range(pop):
            
            if fit[i]==gbest:
                for j in range(n):
                    Lg[j]=L[i][j]
        
        w=wmax-(wmax-wmin)*iteration/itermax;    

        for i in range(pop):
            for j in range(n):
                v[i][j]=(w*v[i][j]+c1*(randint(0,1000)/1000)*(Lp[i][j]-L[i][j])+c2*(randint(0,1000)/1000)*(Lg[j]-L[i][j]))
                if v[i][j]>vmax[j]:
                    v[i][j]=vmax[j]
                elif v[i][j]<=-vmax[j]:
                    v[i][j]=-vmax[j]

        for i in range(pop):
            for j in range(n):
                L[i][j]=L[i][j]+v[i][j]
        global e1
        e1 = (1/gbest)-1;
        if(n==1):
            t.insert(END,str(str(Lg[0])+' '+str(e1)+' '+str(iteration)+'\n'))
            ws.cell(row=iteration+1,column=1).value=str(Lg[0]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=2).value=str(e1).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=3).value=iteration
        elif(n==2):
            t.insert(END,str(str(Lg[0])+' '+str(Lg[1])+' '+str(e1)+' '+str(iteration)+'\n'))
            ws.cell(row=iteration+1,column=1).value=str(str(Lg[0])).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=2).value=str(Lg[1]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=3).value=str(e1).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=4).value=iteration
        elif(n==3):
            t.insert(END,str(str(Lg[0])+' '+str(Lg[1])+' '+str(Lg[2])+' '+str(e1)+' '+str(iteration)+'\n'))
            ws.cell(row=iteration+1,column=1).value=str(Lg[0]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=2).value=str(Lg[1]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=3).value=str(Lg[2]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=4).value=str(e1).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=5).value=iteration
        elif(n==4):
            t.insert(END,str(str(Lg[0])+' '+str(Lg[1])+' '+str(Lg[2])+' '+str(Lg[3])+' '+str(e1)+' '+str(iteration)+'\n'))
            ws.cell(row=iteration+1,column=1).value=str(Lg[0]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=2).value=str(Lg[1]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=3).value=str(Lg[2]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=4).value=str(Lg[3]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=5).value=str(e1).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=6).value=iteration
        elif(n==5):
            t.insert(END,str(str(Lg[0])+' '+str(Lg[1])+' '+str(Lg[2])+' '+str(Lg[3])+' '+str(Lg[4])+' '+str(e1)+' '+str(iteration)+'\n'))
            ws.cell(row=iteration+1,column=1).value=str(Lg[0]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=2).value=str(Lg[1]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=3).value=str(Lg[2]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=4).value=str(Lg[3]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=5).value=str(Lg[4]).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=6).value=str(e1).replace('[','').replace(']','')
            ws.cell(row=iteration+1,column=7).value=iteration
        
        
        if e1<=0:
            print('Converged in iterations ', (iteration))
            savefile()
            showmsg()
            break
        elif iteration==itermax-1:
            print("Not converged ")
        
        window.update_idletasks()
        t.see("end")
        scrollbar.config(command=t.yview)
        scrollbarx.config(command=t.xview)

    
###################################pso###################################




Vars=['x','y','z','r','s']
window =Tk()
window.geometry('500x600')
window.state('zoomed')
window.title("Particle Swarm Optimizer");
window.wm_iconbitmap('icon_ZAw_icon.ico')
var=IntVar()
c1=StringVar()
c2=StringVar()
wmax=StringVar()
wmin=StringVar()
num_p=StringVar()
Excel_file=StringVar()
Excel_file.set("")
lim=[[0,0],[0,0],[0,0],[0,0],[0,0]]
eqn=StringVar()
global scrollbar
scrollbar = Scrollbar(window)
scrollbar.pack(side=RIGHT,fill=Y)
global scrollbarx
scrollbarx = Scrollbar(window)
scrollbarx.pack()
scrollbarx.place(x=905,y=661,width=448,height=40)
scrollbarx.config(orient=HORIZONTAL)
create_gui()
global ws
global wb

def savefile():
    wb.save(str(".\\"+Excel_file.get()));

def get_file_and_opt():
     if(Excel_file.get()==""):
        new('<Ctrl-N>')
     pso()
    
start=Button(window,text="OPTIMIZE",command=get_file_and_opt,font='fixedsys 16')
start.pack();
start.place(x=500,y=600)
t.config(state=DISABLED)
mainloop()


    
